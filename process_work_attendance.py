#! python3
# work_attendance.py 帮助助理整合考勤指纹系统和OA系统的考勤数据，生成每月的最终考勤报表

import openpyxl
from openpyxl.styles import Font,Color,Fill
from openpyxl.styles.colors import RED
import time
import datetime
import logging
import os

logging.basicConfig( filename='myProgramLog.txt', level= logging.DEBUG, format=' %(asctime) s - %(levelname) s - %(message) s')


def processKqTable(processType, totalSheet, inputSheet,totalRowOffset,totalColumnOffset,inputColumnOffset):

    #logging.debug('处理'+processType+' '+totalSheet.cell(row= totalRowOffset, column=2).value+' '+lineDate)
    
    for row_index in range(2, inputSheet.max_row+1):
        if inputSheet.cell(row=row_index, column=2).value == totalSheet.cell(row= totalRowOffset, column=2).value:

            #logging.debug(inputSheet.cell(row=row_index, column=inputColumnOffset).value+ inputSheet.cell(row=row_index, column=inputColumnOffset+1).value)
            timeStart=inputSheet.cell(row=row_index, column=inputColumnOffset).value
            timeStop=inputSheet.cell(row=row_index, column=inputColumnOffset+1).value
                 
            timeStartArr=timeStart.split(' ')
            timeStopArr=timeStop.split(' ')
            # logging.debug(timeStartArr[0],timeStartArr[1],timeStopArr[0],timeStopArr[1])
                 
            startYearMonthDay=timeStartArr[0].split('-')
            stopYearMonthDay=timeStopArr[0].split('-')
            startTime=timeStartArr[1]
            stopTime=timeStopArr[1]
            #print(yearMonthDay,startYearMonthDay,stopYearMonthDay)
            if startYearMonthDay==stopYearMonthDay:#单日
                if yearMonthDay==startYearMonthDay:#当前日期匹配
                    logging.debug('单日'+processType+yearMonthDay[0]+yearMonthDay[1]+yearMonthDay[2]+startYearMonthDay[0]+startYearMonthDay[1]+startYearMonthDay[2])
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset).value = startTime
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+1).value = stopTime
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset).style='Accent2'
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+1).style='Accent2'
                    if processType == '请假':
                         fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+2).value = inputSheet.cell(row=row_index, column=4).value
                         fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+3).value = inputSheet.cell(row=row_index, column=8).value

            elif startYearMonthDay[2]<=yearMonthDay[2] and stopYearMonthDay[2]>=yearMonthDay[2]:#多日而且当前日期在此区间
                if yearMonthDay==startYearMonthDay:#开始日
                    logging.debug(processType+'开始日'+ yearMonthDay[0]+yearMonthDay[1]+yearMonthDay[2]+startYearMonthDay[0]+startYearMonthDay[1]+startYearMonthDay[2])
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset).value = startTime
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+1).value = '18:00:00'    
                elif yearMonthDay==stopYearMonthDay:  #结束日
                    logging.debug(processType+'结束日'+ yearMonthDay[0]+yearMonthDay[1]+yearMonthDay[2]+stopYearMonthDay[0]+stopYearMonthDay[1]+stopYearMonthDay[2])
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset).value = '09:00:00'
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+1).value = stopTime
                else:#进行中
                    logging.debug(processType+'进行中'+ yearMonthDay[0]+yearMonthDay[1]+yearMonthDay[2]+startYearMonthDay[0]+startYearMonthDay[1]+startYearMonthDay[2])
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset).value = '09:00:00'
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+1).value = '18:00:00'
                fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset).style='Accent2'
                fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+1).style='Accent2'
                if processType == '请假':
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+2).value = inputSheet.cell(row=row_index, column=4).value
                    fingerSheet.cell(row= totalRowOffset, column=totalColumnOffset+3).value = inputSheet.cell(row=row_index, column=8).value

            #if inputSheet.cell(row=row_index, column=3).value != totalSheet.cell(row= totalRowOffset, column=3).value:    
                 #logging.debug(inputSheet.cell(row=row_index, column=2).value+'出现在'+inputSheet.cell(row=row_index, column=3).value+'和'+fingerSheet.cell(row= totalRowOffset, column=3).value)
       # else:
       #      print(chuchaiSheet.cell(row=row_chuchai, column=2).value,fingerSheet.cell(row= row_line, column=2).value,'不是一个人')

def checkKqTable(Sheet, curLine):
    time=[Sheet.cell(row= curLine, column=5).value,Sheet.cell(row= curLine, column=6).value,Sheet.cell(row= curLine, column=7).value,Sheet.cell(row= curLine, column=8).value,\
          Sheet.cell(row= curLine, column=11).value,Sheet.cell(row= curLine, column=12).value,Sheet.cell(row= curLine, column=13).value,Sheet.cell(row= curLine, column=14).value]
    
    if Sheet.cell(row= curLine, column=5).value=='invalid' and Sheet.cell(row= curLine, column=7).value==None \
       and Sheet.cell(row= curLine, column=11).value==None and Sheet.cell(row= curLine, column=13).value==None:
        return    
    elif Sheet.cell(row= curLine, column=5).value!='invalid' and Sheet.cell(row= curLine, column=5).value==Sheet.cell(row= curLine, column=6).value \
         and Sheet.cell(row= row_line, column=5).style != 'Accent6':
        print(Sheet.cell(row= row_line, column=2).value, Sheet.cell(row= row_line, column=4).value,'颜色标注缺失')
        print(time,Sheet.cell(row= curLine, column=5).value,Sheet.cell(row= curLine, column=6).value,curLine)
    else:
        #print(time,index)
        officeHour =0
        officeMin=0
        vocationHour =0
        vocationMin=0
        journalHour =0
        journalMin=0
        localjournalHour =0
        localjournalMin=0

        if Sheet.cell(row= curLine, column=5).value!='invalid':#在公司的时间
            leaveTiming = time[1].split(':')
            arriveTiming = time[0].split(':')
        
            if int(leaveTiming[1]) < int(arriveTiming[1]):
              #  print(leaveTiming, arriveTiming)
                officeHour = int(leaveTiming[0])-int(arriveTiming[0])-1
                officeMin = int(leaveTiming[1]) - int(arriveTiming[1])+60
            else:
                officeHour = int(leaveTiming[0])-int(arriveTiming[0])
                officeMin = int(leaveTiming[1]) - int(arriveTiming[1])
            if officeHour < 9:
               # print(str(officeHour),str(officeMin), type(officeHour))
                Sheet.cell(row= row_line, column=1).style = 'Accent4'

                
        if Sheet.cell(row= curLine, column=7).value!= None: #请假的时间
            vocationStop = time[3].split(':')
            vocationStart = time[2].split(':')
      #      print('请假', vocationStop, vocationStart)
            if int(vocationStop[1]) < int(vocationStart[1]):
                vocationHour = int(vocationStop[0])- int(vocationStart[0])-1
                vocationMin = int(vocationStop[1])- int(vocationStart[1])+60
            else:    
                vocationHour = int(vocationStop[0])- int(vocationStart[0])
                vocationMin = int(vocationStop[1])- int(vocationStart[1])

               
        if Sheet.cell(row= curLine, column=11).value!= None: #出差的时间
            journalStop = time[5].split(':')
            journalStart = time[4].split(':')
        #    print('出差',journalStop, journalStart)
            if int(journalStop[1]) < int(journalStart[1]):
                journalHour = int(journalStop[0])- int(journalStart[0])-1
                journalMin = int(journalStop[1])- int(journalStart[1])+60
            else:    
                journalHour = int(journalStop[0])- int(journalStart[0])
                journalMin = int(journalStop[1])- int(journalStart[1])

        if Sheet.cell(row= curLine, column=13).value!= None: #外勤时间
            localjournalStop = time[7].split(':')
            localjournalStart = time[6].split(':')
       #     print('外勤',localjournalStop, localjournalStart)
            if int(localjournalStop[1]) < int(localjournalStart[1]):
                localjournalHour = int(localjournalStop[0])- int(localjournalStart[0])-1
                localjournalMin = int(localjournalStop[1])- int(localjournalStart[1])+60
            else:    
                localjournalHour = int(localjournalStop[0])- int(localjournalStart[0])
                localjournalMin = int(localjournalStop[1])- int(localjournalStart[1])

        if localjournalMin+journalMin+vocationMin+officeMin > 60:
            dutyHour = localjournalHour+journalHour+vocationHour+officeHour+1
        else:
            dutyHour = localjournalHour+journalHour+vocationHour+officeHour

        if dutyHour < 9:
            #print(Sheet.cell(row= row_line, column=2).value, Sheet.cell(row= row_line, column=4).value,curLine)
            #print(str(localjournalHour),str(journalHour),str(vocationHour),str(officeHour))
            #print(str(localjournalMin),str(journalMin),str(vocationMin),str(officeMin))
            Sheet.cell(row= row_line, column=2).style = 'Accent4'

            
#logging.debug('读取月度考勤原始数据表,然后处理外勤，请假，加班和出差，同时会标注考勤时间异常，帮助检查是否有补卡')


currentDir=os.getcwd()
sheetList=os.listdir(currentDir)
fmark = 0
wmark = 0
qmark = 0
jmark = 0
cmark = 0
for peformanceSheet in sheetList:
    sheetName=peformanceSheet.split('.')
    #print(peformanceSheet)
    if sheetName[1] == 'xlsx'and sheetName[0]!='当月考勤原始数据汇总':
        tableOpen=openpyxl.load_workbook(str(currentDir+'\\'+peformanceSheet),data_only=True)
        allSheetNames=tableOpen.get_sheet_names()
        #print(allSheetNames)
        logging.debug(allSheetNames)
        for curSheet in allSheetNames:
            if '考勤机' in curSheet:
                fingerSheet=tableOpen.get_sheet_by_name(curSheet)
                finger_kq=tableOpen
                fmark+=1    
            if '外勤' in curSheet:
                print('外勤sheet名字匹配成功')
                waiqinSheet=tableOpen.get_sheet_by_name(curSheet)
                wmark+=1
            elif '请假' in curSheet:
                print('请假sheet名字匹配成功')
                qingjiaSheet=tableOpen.get_sheet_by_name(curSheet)
                qmark+=1
            elif '加班' in curSheet:
                print('加班sheet名字匹配成功')
                jiabanSheet=tableOpen.get_sheet_by_name(curSheet)
                jmark+=1
            elif '出差' in curSheet:
                print('出差sheet名字匹配成功')
                chuchaiSheet=tableOpen.get_sheet_by_name(curSheet)
                cmark+=1


logging.debug(fingerSheet.title)
logging.debug(waiqinSheet.title)
logging.debug(chuchaiSheet.title)
logging.debug(qingjiaSheet.title)
logging.debug(jiabanSheet.title)

logging.debug('首先合并所有信息到一个总表, 该表格目前有'+ str(fingerSheet.max_row)+'行要处理')
fingerSheet['F1'] = '离开公司'+ str(fingerSheet['E1'].value)
fingerSheet['E1'] = '进入公司'+ str(fingerSheet['E1'].value)

ft = Font(bold=True)

fingerSheet['G1']= '请假开始时间'
fingerSheet['G1'].font=ft

fingerSheet['H1']= '请假结束时间'
fingerSheet['H1'].font=ft

fingerSheet['I1']= '请假类型'
fingerSheet['I1'].font=ft

fingerSheet['J1']= '请假原因'
fingerSheet['J1'].font=ft

fingerSheet['K1']= '出差开始时间'
fingerSheet['K1'].font=ft

fingerSheet['L1']= '出差结束时间'
fingerSheet['L1'].font=ft

fingerSheet['M1']= '外勤开始时间'
fingerSheet['M1'].font=ft

fingerSheet['N1']= '外勤结束时间'
fingerSheet['N1'].font=ft

fingerSheet['O1']= '加班开始时间'
fingerSheet['O1'].font=ft

fingerSheet['P1']= '加班结束时间'
fingerSheet['P1'].font=ft

fingerSheet['Q1']= '补卡开始时间'
fingerSheet['Q1'].font=ft

fingerSheet['R1']= '补卡结束时间'
fingerSheet['R1'].font=ft

fingerSheet.freeze_panes = 'F2'


for row_line in range(2,fingerSheet.max_row+1):     #skip the first line
   
  #周末的日期highlight显示，先处理，后面周末加班时的单次考勤要更换highlight背景。
  #获取该行的日期
    lineDate=fingerSheet.cell(row= row_line, column=4).value
    yearMonthDay=lineDate.split('-')
    weekday = datetime.date(int(yearMonthDay[0]), int(yearMonthDay[1]), int(yearMonthDay[2])).weekday()

    if weekday == 5 or weekday == 6:
        fingerSheet.cell(row= row_line, column=5).style = 'Accent3'
        fingerSheet.cell(row= row_line, column=6).style = 'Accent3'

 #拆分考勤时间  
    produceName = fingerSheet.cell(row= row_line, column=5 ).value
    if produceName=='':
        fingerSheet.cell(row= row_line, column=5).value = 'invalid'
        fingerSheet.cell(row= row_line, column=6).value = 'invalid'
        #print(row_line, type(produceName))
    else:
        strTimeOfFingerprint=produceName.split(' ')
        #只有一次指纹记录的
        if len(strTimeOfFingerprint) == 1 or strTimeOfFingerprint[0]==strTimeOfFingerprint[-1]:
           # print(row_line, produceName,strTimeOfFingerprint,strTimeOfFingerprint[0],strTimeOfFingerprint[-1],len(strTimeOfFingerprint))
            fingerSheet.cell(row= row_line, column=5).style = 'Accent6'
            fingerSheet.cell(row= row_line, column=6).style = 'Accent6'
        fingerSheet.cell(row= row_line, column=5).value = strTimeOfFingerprint[0]
        fingerSheet.cell(row= row_line, column=6).value = strTimeOfFingerprint[-1]
  #处理表格
  #def processKqTable(processType, totalSheet, inputSheet,totalRowOffset,totalColumnOffset,inputColumnOffset):        
    processKqTable('请假', fingerSheet, qingjiaSheet,row_line,7,5)
    processKqTable('出差', fingerSheet, chuchaiSheet,row_line,11,4)
    processKqTable('外勤', fingerSheet, waiqinSheet,row_line,13,4)
    processKqTable('加班', fingerSheet, jiabanSheet,row_line,15,5)
    if weekday !=5 and weekday !=6:
        checkKqTable(fingerSheet, row_line)
#tableNew=openpyxl.Workbook()
#sheet = tableNew.active
#sheet = fingerSheet
#tableNew.save(str(currentDir+'\\'+'result.xlsx'))

finger_kq.save(str(currentDir+'\\'+'当月考勤原始数据汇总.xlsx'))




